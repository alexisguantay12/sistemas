# presupuestos/views.py
# ============================================================
# 📦 Importaciones
# ============================================================

from django.shortcuts import render, redirect, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib import messages
from django.db import transaction
from decimal import Decimal, InvalidOperation
import json
import os
import locale
import io

from openpyxl import load_workbook
from reportlab.lib.pagesizes import A4
from reportlab.lib import colors
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.platypus import Table, TableStyle
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.utils import ImageReader
from django.contrib.auth.decorators import login_required
from django.views.decorators.cache import never_cache
from .models import (
    Presupuesto, PresupuestoItem, Prestacion, PresupuestoHistorial, Pago, Medico, ObraSocial,Reintegro
)
from .forms import NomencladorUploadForm
from django.db.models import Q
from django.db.models import Case, When, Value, IntegerField
# ============================================================
# 🧾 Listado de Presupuestos
# ============================================================
@login_required
def lista_presupuestos(request):
    presupuestos = Presupuesto.objects.all().order_by('-fecha_creacion')
    return render(request, 'presupuestos/presupuestos.html', {'presupuestos': presupuestos})











# ============================================================
# 💰 Registro de Pagos
# ============================================================

from django.template.loader import render_to_string
from django.utils import timezone
from django.http import JsonResponse


def registrar_pago(request, pk):
    presupuesto = get_object_or_404(Presupuesto, id=pk)

    if request.method == 'POST' and request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        monto = request.POST.get('monto')
        medio = request.POST.get('medio_pago')
        comentario = request.POST.get('observaciones_pago')
        fecha_str = request.POST.get("fecha")
                    # 2️⃣ Si el usuario ingresó fecha, combinarla con la hora actual

        print("Pago come",comentario)
        if fecha_str:
            fecha_base = datetime.strptime(fecha_str, "%Y-%m-%d")  # convierte el string a fecha
            hora_actual = datetime.now().time()                   # hora actual del sistema
            fecha = datetime.combine(fecha_base, hora_actual)
        else:
            fecha = datetime.now()  # si no hay fecha, usar fecha/hora actuales
        
        if not monto:
            return JsonResponse({'success': False, 'error': 'Monto inválido'})

        try:
            with transaction.atomic():
                # Crear el pago
                Pago.objects.create(
                    presupuesto=presupuesto,
                    monto=monto,
                    fecha=fecha,
                    medio_pago=medio,
                    observaciones=comentario,
                    user_made=request.user
                )

                # Cambiar estado del presupuesto
                if presupuesto.estado != "en_curso":
                    presupuesto.estado = "en_curso"
                    presupuesto.save()

            # Actualizar listado de pagos
            pagos = presupuesto.pagos.all().order_by('-fecha')
            for p in pagos:
                p.puede_eliminar = (
                    (timezone.now() - p.created_at).total_seconds() < 900 and p.user_made == request.user
                ) or request.user.groups.filter(name='administrador').exists()

            pagos_html = render_to_string('presupuestos/partials/_tabla_pagos.html', {'pagos': pagos})
            total_pagado = sum(pago.monto for pago in presupuesto.pagos.all())
            saldo = presupuesto.saldo


            pagos = presupuesto.pagos.all().order_by("-fecha", "-id")
            reintegros = presupuesto.reintegros.all().order_by("-fecha", "-id")

            resumen_movimientos_html = render_to_string(
                "presupuestos/partials/_resumen_movimientos.html",
                {
                    "pagos": pagos,
                    "reintegros": reintegros,
                    "saldo": saldo,
                },
                request=request
            )

            return JsonResponse({
                'success': True,
                'pagos_html': pagos_html,
                'resumen_movimientos_html':resumen_movimientos_html,
                'total_pagado': total_pagado,
                'saldo': float(saldo),
                'estado':presupuesto.get_estado_display()
            })

        except Exception as e:
            return JsonResponse({'success': False, 'error': f'Error al registrar el pago: {str(e)}'})

    return JsonResponse({'success': False, 'error': 'Petición no válida'})
# ============================================================
# 📜 Detalle de Presupuesto + Historial + Pagos
# ============================================================
from django.db.models import Sum
from datetime import datetime, time
def registrar_reintegro(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)

    if request.method == "POST":
        try:
            monto = Decimal(request.POST.get("monto", "0"))
            medio_pago = request.POST.get("medio_pago")
            observaciones = request.POST.get("observaciones_reintegro", "")
            fecha_str = request.POST.get("fecha")
 

            # 2️⃣ Si el usuario ingresó fecha, combinarla con la hora actual
            if fecha_str:
                fecha_base = datetime.strptime(fecha_str, "%Y-%m-%d")  # convierte el string a fecha
                hora_actual = datetime.now().time()                   # hora actual del sistema
                fecha = datetime.combine(fecha_base, hora_actual)
            else:
                fecha = datetime.now()  # si no hay fecha, usar fecha/hora actuales
 
            print(fecha)

            reintegro = Reintegro.objects.create(
                presupuesto=presupuesto,
                monto=monto,
                medio_pago=medio_pago,
                observaciones=observaciones,
                fecha=fecha,
                user_made=request.user
            )
            print("Si pase lo anterior ")
            reintegros = presupuesto.reintegros.all().order_by("-fecha", "-id")
            total_reintegrado = reintegros.aggregate(total=Sum("monto"))["total"] or Decimal("0") 
            for r in reintegros:
                r.puede_eliminar = (
                    (timezone.now() - r.created_at).total_seconds() < 900 and r.user_made == request.user
                ) or request.user.groups.filter(name='administrador').exists()

            html = render_to_string(
                "presupuestos/partials/_tabla_reintegros.html",
                {"reintegros": reintegros},
                request=request
            )
            
            
            pagos = presupuesto.pagos.all().order_by("-fecha", "-id") 
        
            resumen_movimientos_html = render_to_string(
                "presupuestos/partials/_resumen_movimientos.html",
                {
                    "pagos": pagos,
                    "reintegros": reintegros,
                    "saldo": presupuesto.saldo,
                },
                request=request
            )

            return JsonResponse({
                "success": True,
                "reintegros_html": html,
                "resumen_movimientos_html": resumen_movimientos_html,
                "total_reintegrado": float(total_reintegrado),
                "saldo": float(presupuesto.saldo),
                "estado": presupuesto.get_estado_display(), 
            })
            
        except Exception as e:
            return JsonResponse({"success": False, "error": str(e)})

    return JsonResponse({"success": False, "error": "Método no permitido"})



from django.views.decorators.http import require_POST

@require_POST
def eliminar_pago(request, pk):
    pago = get_object_or_404(Pago, id=pk)
    presupuesto = pago.presupuesto

    try:
        with transaction.atomic():
            # Marcar usuario que eliminó y borrar el pago
            pago.user_deleted = request.user
            pago.delete()

            # Si no quedan pagos, actualizar estado del presupuesto
            if not presupuesto.pagos.exists():
                presupuesto.estado = "autorizado"
                presupuesto.save()

        # Actualizar datos de respuesta
        pagos = presupuesto.pagos.all().order_by('-fecha')
        for p in pagos:
            p.puede_eliminar = (
                (timezone.now() - p.fecha).total_seconds() < 900 and p.user_made == request.user
            ) or request.user.groups.filter(name='administrador').exists()

        total_pagado = sum(pago.monto for pago in presupuesto.pagos.all())
        saldo = presupuesto.saldo

        pagos_html = render_to_string('presupuestos/partials/_tabla_pagos.html', {'pagos': pagos})

        pagos = presupuesto.pagos.all().order_by("-fecha", "-id")
        reintegros = presupuesto.reintegros.all().order_by("-fecha", "-id")

        resumen_movimientos_html = render_to_string(
            "presupuestos/partials/_resumen_movimientos.html",
            {
                "pagos": pagos,
                "reintegros": reintegros,
                "saldo": presupuesto.saldo,
            },
            request=request
        )


        return JsonResponse({
            'success': True,
            'pagos_html': pagos_html,
            'resumen_movimiento_html':resumen_movimientos_html,
            'total_pagado': total_pagado,
            'saldo': float(saldo),
            'estado':presupuesto.get_estado_display()
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'})



@require_POST
def eliminar_reintegro(request, pk):
    reintegro = get_object_or_404(Reintegro, id=pk)
    presupuesto = reintegro.presupuesto

    try:
        with transaction.atomic():
            # Marcar usuario que eliminó y borrar el pago
            reintegro.user_deleted = request.user
            reintegro.delete() 

        # Actualizar datos de respuesta
        reintegros = presupuesto.reintegros.all().order_by('-fecha')
        for r in reintegros:
            r.puede_eliminar = (
                (timezone.now() - r.fecha).total_seconds() < 900 and r.user_made == request.user
            ) or request.user.groups.filter(name='administrador').exists()

        total_reintegrado = sum(reintegro.monto for reintegro in presupuesto.reintegros.all())
        saldo = presupuesto.saldo

        reintegros_html = render_to_string('presupuestos/partials/_tabla_reintegros.html', {'reintegros': reintegros})

 

        pagos = presupuesto.pagos.all().order_by("-fecha", "-id")
        reintegros = presupuesto.reintegros.all().order_by("-fecha", "-id")

        resumen_movimientos_html = render_to_string(
            "presupuestos/partials/_resumen_movimientos.html",
            {
                "pagos": pagos,
                "reintegros": reintegros,
                "saldo": presupuesto.saldo,
            },
            request=request
        )


        return JsonResponse({
            'success': True,
            'reintegros_html': reintegros_html,
            'resumen_movimientos_html':resumen_movimientos_html,
            'total_reintegrado': total_reintegrado,
            'saldo': float(saldo),
            'estado':presupuesto.get_estado_display()
        })

    except Exception as e:
        return JsonResponse({'success': False, 'error': f'Error al eliminar el pago: {str(e)}'})


@login_required
def guardar_datos_internacion(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)

    if request.method == "POST":
        try:
            episodio = request.POST.get("episodio")
            hc = request.POST.get("hc")
            fecha_inicio_str = request.POST.get("fecha_inicio")
            fecha_fin_str = request.POST.get("fecha_fin")

            presupuesto.episodio = episodio or None
            presupuesto.hc = hc or None

            presupuesto.fecha_inicio = (
                timezone.make_aware(datetime.strptime(fecha_inicio_str, "%Y-%m-%d"))
                if fecha_inicio_str else None
            )

            presupuesto.fecha_fin = (
                timezone.make_aware(datetime.strptime(fecha_fin_str, "%Y-%m-%d"))
                if fecha_fin_str else None
            )

            presupuesto.user_updated = request.user
            presupuesto.save()

            return JsonResponse({
                "success": True,
                "episodio": presupuesto.episodio,
                "hc": presupuesto.hc,
                "fecha_inicio": presupuesto.fecha_inicio.strftime("%d/%m/%Y") if presupuesto.fecha_inicio else "-",
                "fecha_fin": presupuesto.fecha_fin.strftime("%d/%m/%Y") if presupuesto.fecha_fin else "-"
            })

        except Exception as e:
            return JsonResponse({
                "success": False,
                "error": str(e)
            })

    return JsonResponse({
        "success": False,
        "error": "Método no permitido"
    })


@require_POST
def cerrar_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, id=pk)

    if presupuesto.estado == "cerrado":
        return JsonResponse({
            "success": False,
            "error": "El presupuesto ya está cerrado."
        })

    try:
        with transaction.atomic():
            presupuesto.estado = "cerrado"
            presupuesto.save()
        return JsonResponse({
            "success": True,
            "nuevo_estado": presupuesto.get_estado_display()
        })
    except Exception as e:
        return JsonResponse({
            "success": False,
            "error": f"Error al cerrar el presupuesto: {e}"
        })


@login_required
def detalle_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    historial = presupuesto.historiales.all().order_by('fecha')
    pagos = presupuesto.pagos.all().order_by('-fecha')
    reintegros = presupuesto.reintegros.all().order_by('-fecha')
    total_pagado = sum(pago.monto for pago in presupuesto.pagos.all())
    total_reintegrado = sum(reintegro.monto for reintegro in presupuesto.reintegros.all())
    saldo = presupuesto.saldo
    for p in pagos:
        p.puede_eliminar = (((timezone.now() - p.fecha).total_seconds() < 900 and p.user_made==request.user) or request.user.groups.filter(name='administrador').exists()) and presupuesto.estado!='cerrado'
    for r in reintegros:
        r.puede_eliminar = (((timezone.now()-r.created_at).total_seconds()< 900 and r.user_made == request.user) or request.user.groups.filter(name='administrados').exists()) and presupuesto.estado != 'cerrado'
 
    historial_json = []
    for h in historial:
        datos = h.datos  # dict con todos los datos guardados
        historial_json.append({
            'id': h.id,
            'fecha': h.fecha.strftime('%d/%m/%Y %H:%M'),
            'user': f"{h.user_made.first_name} {h.user_made.last_name}" if h.user_made else "Desconocido",
            'paciente': {
                'nombre': datos['paciente'].get('nombre', ''),
                'dni': datos['paciente'].get('dni', ''),
                'edad': datos['paciente'].get('edad', ''),
                'telefono': datos['paciente'].get('telefono', ''),
                'email': datos['paciente'].get('email', ''),
                'obra_social': datos['paciente'].get('obra_social', ''),
            },
            'medico': datos.get('medico', ''),
            'diagnostico': datos.get('diagnostico', ''),
            'estado': datos.get('estado', ''),
            'motivo_no_concretado': datos.get('motivo_no_concretado', ''),
            'items': [
                {
                    'codigo': it.get('codigo', ''),
                    'tipo': it.get('tipo', ''),
                    'prestacion': it.get('prestacion', ''),
                    'cantidad': float(it.get('cantidad', 0)),
                    'precio': float(it.get('precio', 0)),
                    'importe': float(it.get('importe', 0)),
                    'iva': float(it.get('iva', 0)),
                    'subtotal': float(it.get('subtotal', 0)),
                } for it in datos.get('items', [])
            ],
            'total': float(datos.get('total', 0)),
        })

    

    itemspresupuesto = (
        presupuesto.items
        .annotate(
            prioridad=Case(
                When(codigo__regex=r'^\d+$', then=Value(0)),            # si el código es numérico
                When(codigo__istartswith="med", then=Value(2)),         # si empieza con "med"
                default=Value(1),                                       # el resto
                output_field=IntegerField(),
            )
        )
        .order_by("prioridad", "codigo")  # ordena primero por prioridad, luego por código
    ) 
    context = {
        'presupuesto': presupuesto,
        'today': timezone.now().date(),
        'itemspresupuesto':itemspresupuesto,
        'historial_json': json.dumps(historial_json),
        'pagos': pagos,
        'reintegros': reintegros,
        'total_pagado':total_pagado,
        'total_reintegrado':total_reintegrado,
        'saldo':saldo
    } 
    return render(request, 'presupuestos/detalle_presupuesto.html', context)


# ============================================================
# 🕒 Guardar Historial (helper)
# ============================================================

def guardar_historial(presupuesto, usuario=None):
    """Guarda una copia completa del estado actual del presupuesto."""

    itemspresupuesto = (
        presupuesto.items
        .annotate(
            prioridad=Case(
                When(codigo__regex=r'^\d+$', then=Value(0)),            # si el código es numérico
                When(codigo__istartswith="med", then=Value(2)),         # si empieza con "med"
                default=Value(1),                                       # el resto
                output_field=IntegerField(),
            )
        )
        .order_by("prioridad", "codigo")  # ordena primero por prioridad, luego por código
    )

    items = [{
        "codigo": item.codigo,
        "tipo": item.get_tipo_display(),
        "prestacion": item.prestacion,
        "cantidad": float(item.cantidad),
        "precio": float(item.precio),
        "importe": float(item.importe),
        "iva": float(item.iva),
        "subtotal": float(item.subtotal),
    } for item in itemspresupuesto]

    datos = {
        "paciente": {
            "nombre": presupuesto.paciente_nombre,
            "dni": presupuesto.paciente_dni,
            "edad": presupuesto.paciente_edad,
            "direccion": presupuesto.paciente_direccion,
            "telefono": presupuesto.paciente_telefono,
            "email": presupuesto.paciente_email,
            "obra_social": presupuesto.obra_social.nombre,
        },
        "medico": presupuesto.medico.nombre,
        "diagnostico": presupuesto.diagnostico,
        "estado": presupuesto.estado,
        "motivo_no_concretado": presupuesto.motivo_no_concretado,
        "items": items,
        "total": float(presupuesto.total),
    }

    PresupuestoHistorial.objects.create(
        presupuesto=presupuesto,
        user_made=usuario,
        datos=datos
    )


# ============================================================
# 📖 Detalle del Historial
# ============================================================

def detalle_historial(request, historial_id):
    historial = get_object_or_404(PresupuestoHistorial, id=historial_id)
    return render(request, "presupuestos/detalle_historial.html", {"historial": historial})

def autorizar_presupuesto(request, pk):
    if request.method != "POST":
        return JsonResponse({"success": False, "error": "Método no permitido"}, status=405)

    presupuesto = get_object_or_404(Presupuesto, id=pk)

    if presupuesto.estado == "pendiente":
        # ✅ Si tiene pagos, pasa a "en curso"
        if presupuesto.pagos.exists():
            presupuesto.estado = "en_curso"
        else:
            presupuesto.estado = "autorizado"

        presupuesto.save()
        estado_nuevo = presupuesto.get_estado_display()
        return JsonResponse({"success": True, "nuevo_estado": estado_nuevo})
    else:
        return JsonResponse({
            "success": False,
            "error": f"No se puede autorizar un presupuesto con estado '{presupuesto.get_estado_display()}'."
        })
# ============================================================
# ✏️ Edición de Presupuesto
# ============================================================
@never_cache
@login_required
def editar_presupuesto(request, pk):
    presupuesto = get_object_or_404(Presupuesto, pk=pk)
    tiene_iva = presupuesto.items.filter(iva__gt=0).exists()
    print("Entre aqui")
        # 🚫 Bloquear acceso a edición según estado
    if presupuesto.estado in ["cerrado", "expirado"]:
        messages.error(request, "No se puede editar un presupuesto cerrado o expirado.")
        return redirect("presupuestos_app:detalle_presupuesto", pk=presupuesto.pk)
    
    if request.method == "POST":
        # Guardar historial antes de modificar
        try:
            with transaction.atomic():
                guardar_historial(presupuesto, usuario=presupuesto.user_updated) 
                # Actualizar campos principales 
                presupuesto.paciente_nombre = request.POST.get("paciente_nombre") 
                presupuesto.paciente_edad = request.POST.get("paciente_edad") or None
                presupuesto.paciente_direccion = request.POST.get("paciente_direccion")
                presupuesto.paciente_telefono = request.POST.get("paciente_telefono")
                presupuesto.paciente_email = request.POST.get("paciente_email") 
                presupuesto.diagnostico = request.POST.get("diagnostico")
                presupuesto.motivo_no_concretado = request.POST.get("observaciones")
                presupuesto.episodio = request.POST.get("episodio")
                presupuesto.estado = 'pendiente'
                presupuesto.user_updated = request.user
                presupuesto.save()
                print("Entre aqui")
                # Eliminar items anteriores
                presupuesto.items.all().delete()

                # Crear los nuevos desde el POST
                codigos = request.POST.getlist("codigo")
                tipos = request.POST.getlist("tipo")
                prestaciones = request.POST.getlist("prestacion")
                cantidades = request.POST.getlist("cantidad")
                precios = request.POST.getlist("precio")
                importes = request.POST.getlist("importe")
                ivas = request.POST.getlist("iva")
                subtotales = request.POST.getlist("subtotal")
                comentarios = request.POST.getlist("comentario")
                for i in range(len(codigos)):
                    if codigos[i].strip() == "" and prestaciones[i].strip() == "":
                        continue

                    PresupuestoItem.objects.create(
                        presupuesto=presupuesto,
                        codigo=codigos[i] or "",
                        tipo=tipos[i] or "",
                        prestacion=prestaciones[i],
                        cantidad=int(cantidades[i]) if cantidades[i] else 1,
                        precio=float(precios[i].replace(',', '.')) if precios[i] else 0,
                        importe=float(importes[i].replace('.', '').replace(',', '.')) if importes[i] else 0,
                        iva=float(ivas[i].replace('.', '').replace(',', '.')) if ivas[i] else 0,
                        subtotal=float(subtotales[i].replace('.', '').replace(',', '.')) if subtotales[i] else 0,
                        comentario=comentarios[i] or "",
                    )

                messages.success(request, "Presupuesto actualizado correctamente")
                return redirect("presupuestos_app:detalle_presupuesto", pk=presupuesto.pk)
        except Exception as e:
            # Manejo de errores
            print("Error al editar presupuesto y pago:", e)
            return None, None
        
    #Esta parte del codigo ordena por codigo antes de presentarlo en la tabla de presupuesto

    from django.db.models import Case, When, Value, IntegerField    
    itemspresupuesto = (
        presupuesto.items
        .annotate(
            prioridad=Case(
                When(codigo__regex=r'^\d+$', then=Value(0)),            # si el código es numérico
                When(codigo__istartswith="med", then=Value(2)),         # si empieza con "med"
                default=Value(1),                                       # el resto
                output_field=IntegerField(),
            )
        )
        .order_by("prioridad", "codigo")  # ordena primero por prioridad, luego por código
    ) 
    return render(request, "presupuestos/editar_presupuesto.html", {
        "itemspresupuesto":itemspresupuesto,
        "presupuesto": presupuesto,
        "tiene_iva": tiene_iva
    })


import uuid
from datetime import datetime
from django.views.decorators.cache import never_cache

# ============================================================
# ➕ Agregar Presupuesto Nuevo
# ============================================================
@login_required
@never_cache
def agregar_presupuesto(request):
    if request.method == "POST":
        form_token = request.POST.get("form_token")

        # 🛡️ 1) Validar que venga token
        if not form_token:
            messages.error(request, "Formulario inválido.")
            return redirect("presupuestos_app:presupuestos")

        # 🛡️ 2) Ver si el token sigue activo (no usado)
        if not request.session.get(f"form_token_{form_token}", False):
            messages.warning(
                request,
                "Este formulario ya fue enviado. No se procesó nuevamente."
            )
            return redirect("presupuestos_app:presupuestos")

        # 🔒 3) Marcar token como usado
        request.session[f"form_token_{form_token}"] = False

        try:
            with transaction.atomic():
                medico_id = request.POST.get("medico")
                obra_social_id = request.POST.get("obra_social")

                # 1️⃣ Recuperar el valor del input <input type="date" name="fecha_presupuesto">
                fecha_str = request.POST.get("fecha_presupuesto")  # ejemplo: "2025-10-17"

                # 2️⃣ Si el usuario ingresó fecha, combinarla con la hora actual
                if fecha_str:
                    fecha_base = datetime.strptime(fecha_str, "%Y-%m-%d")  # convierte el string a fecha
                    hora_actual = datetime.now().time()                   # hora actual del sistema
                    fecha_final = datetime.combine(fecha_base, hora_actual)
                else:
                    fecha_final = datetime.now()  # si no hay fecha, usar fecha/hora actuales

                # Validar que se haya elegido un mécdico válido
                medico = get_object_or_404(Medico, id=medico_id)
                obra_social = get_object_or_404(ObraSocial, id=obra_social_id)

                presupuesto = Presupuesto.objects.create(
                    paciente_nombre=request.POST.get("paciente_nombre"),
                    paciente_dni=request.POST.get("paciente_dni"),
                    paciente_edad=request.POST.get("paciente_edad") or None,
                    paciente_direccion=request.POST.get("paciente_direccion"),
                    paciente_telefono=request.POST.get("paciente_telefono"),
                    paciente_email=request.POST.get("paciente_email"),
                    obra_social=obra_social,
                    medico=medico,
                    diagnostico=request.POST.get("diagnostico"),
                    motivo_no_concretado=request.POST.get("observaciones"),
                    episodio=request.POST.get("episodio"),
                    fecha_creacion=fecha_final,
                    user_made=request.user,
                    user_updated=request.user
                )

                # Carga de items
                codigos = request.POST.getlist("codigo")
                tipos = request.POST.getlist("tipo")
                prestaciones = request.POST.getlist("prestacion")
                cantidades = request.POST.getlist("cantidad")
                precios = request.POST.getlist("precio")
                importes = request.POST.getlist("importe")
                ivas = request.POST.getlist("iva")
                subtotales = request.POST.getlist("subtotal")
                comentarios = request.POST.getlist("comentario")

                for i in range(len(prestaciones)):
                    if not prestaciones[i].strip():
                        continue
                    PresupuestoItem.objects.create(
                        presupuesto=presupuesto,
                        codigo=codigos[i] or "",
                        tipo=tipos[i] or "",
                        prestacion=prestaciones[i],
                        cantidad=int(cantidades[i]) if cantidades[i] else 1,
                        precio=float(precios[i].replace(',', '.')) if precios[i] else 0,
                        importe=float(importes[i].replace('.', '').replace(',', '.')) if importes[i] else 0,
                        iva=float(ivas[i].replace('.', '').replace(',', '.')) if ivas[i] else 0,
                        subtotal=float(subtotales[i].replace('.', '').replace(',', '.')) if subtotales[i] else 0,
                        comentario=comentarios[i] or "",
                    )

                return redirect("presupuestos_app:presupuestos")
        except Exception as e:
            print("Error al crear presupuesto:", e)
            messages.error(request, "Ocurrió un error al crear el presupuesto.")
            return redirect("presupuestos_app:presupuestos")

    # GET → generar token nuevo
    form_token = str(uuid.uuid4())
    request.session[f"form_token_{form_token}"] = True

    prestaciones = Prestacion.objects.all()
    medicos = Medico.objects.all().order_by('nombre')
    obras_sociales = ObraSocial.objects.all().order_by('nombre')

    return render(
        request,
        "presupuestos/agregar_presupuesto.html",
        {
            "obras_sociales": obras_sociales,
            "prestaciones": prestaciones,
            "medicos": medicos,
            "form_token": form_token,
        }
    )


from django.shortcuts import render, get_object_or_404
from django.db.models import Q
from django.contrib.auth.decorators import login_required

from .models import ObraSocial

@login_required
def gestion_clausulas(request):
    mensaje_clausula = None
    error_clausula = None

    if request.method == "POST":
        accion = request.POST.get("accion")
        obra_id = request.POST.get("obra_id_edit")
        porcentaje = request.POST.get("porcentaje")
        obra_select = request.POST.get("obra_social")

        # -------------------------
        # BORRAR (poner porcentaje en 0)
        # -------------------------
        if accion == "borrar":
            obra = get_object_or_404(ObraSocial, id=obra_id)
            obra.porcentaje = 0
            obra.save()
            mensaje_clausula = f"Se eliminó la cláusula de {obra.nombre}."
        
        # -------------------------
        # GUARDAR / EDITAR
        # -------------------------
        elif accion == "guardar":
            try:
                porcentaje = int(porcentaje)
            except:
                error_clausula = "El porcentaje debe ser un número entero."
            else:
                if porcentaje < 0 or porcentaje > 100:
                    error_clausula = "El porcentaje debe estar entre 0 y 100."

            if not error_clausula:
                if obra_id:  # editar
                    obra = get_object_or_404(ObraSocial, id=obra_id)
                else:  # crear
                    obra = get_object_or_404(ObraSocial, id=obra_select)

                obra.porcentaje = porcentaje
                obra.save()
                mensaje_clausula = f"Cláusula guardada para {obra.nombre}."

    obras_con_clausula = ObraSocial.objects.filter(porcentaje__gt=0)
    obras_sin_clausula = ObraSocial.objects.filter(
        Q(porcentaje__isnull=True) | Q(porcentaje=0)
    ).order_by("nombre")

    return render(request, "presupuestos/clausulas.html", {
        "obras_con_clausula": obras_con_clausula,
        "obras_sin_clausula": obras_sin_clausula,
        "mensaje_clausula": mensaje_clausula,
        "error_clausula": error_clausula,
    })


# ============================================================
# 🔎 Buscar prestación por código
# ============================================================

def get_prestacion(request, codigo):
    """Devuelve los datos de una prestación según código."""
    prestacion = get_object_or_404(Prestacion, codigo=codigo)

    # Casos especiales
    if codigo == '430101':
        p_desc = get_object_or_404(Prestacion, codigo='431001')
        precio_u = (prestacion.gastos + prestacion.especialista) + (p_desc.gastos + p_desc.especialista)
    elif codigo == '400101':
        p_desc = get_object_or_404(Prestacion, codigo='431002')
        precio_u = (prestacion.gastos + prestacion.especialista) + (p_desc.gastos + p_desc.especialista)
    elif codigo == '430130':
        p_desc = get_object_or_404(Prestacion, codigo='431001')
        precio_u = (prestacion.gastos + prestacion.especialista) + (p_desc.gastos + p_desc.especialista)
    elif codigo == '340907':
        precio_u = (prestacion.gastos + prestacion.especialista) * 3
    else:
        precio_u = prestacion.gastos

    obra_id = request.GET.get("obra_social")
    # aplicar cláusula
    if obra_id:
        try:
            obra = ObraSocial.objects.get(id=obra_id)
            if obra.porcentaje and obra.porcentaje > 0:
                precio_u = precio_u + (precio_u * obra.porcentaje / 100)
        except:
            pass
    data = {
        "prestacion": prestacion.codigo,
        "nombre": prestacion.nombre,
        "precio": float(precio_u),
        "tipo": prestacion.tipo or "",
    }
    return JsonResponse(data)


# ============================================================
# 💡 Obtener precio según tipo (gastos, especialista, total)
# ============================================================

def get_tipo(request, codigo):
    tipo = request.GET.get("tipo")
    try:
        prestacion = Prestacion.objects.get(codigo=codigo) 
        if tipo == "gastos":
            precio = prestacion.gastos
        elif tipo == "especialista":
            precio = prestacion.especialista
        else:
            precio = prestacion.total()

        # Casos especiales
        if codigo == '430101':
            p_desc = get_object_or_404(Prestacion, codigo='431001')
            precio = (prestacion.gastos + prestacion.especialista) + (p_desc.gastos + p_desc.especialista)
        elif codigo == '400101':
            p_desc = get_object_or_404(Prestacion, codigo='431002')
            precio = (prestacion.gastos + prestacion.especialista) + (p_desc.gastos + p_desc.especialista)
        elif codigo == '430130':
            p_desc = get_object_or_404(Prestacion, codigo='431001')
            precio = (prestacion.gastos + prestacion.especialista) + (p_desc.gastos + p_desc.especialista)
        elif codigo == '340907':
            precio = (prestacion.gastos + prestacion.especialista) * 3 
    
    
    except Prestacion.DoesNotExist:
        return JsonResponse({"error": "Código no encontrado"}, status=404)
    obra_id = request.GET.get("obra_social")
    # aplicar cláusula
    if obra_id:
        try:
            obra = ObraSocial.objects.get(id=obra_id)
            if obra.porcentaje and obra.porcentaje > 0:
                precio = precio + (precio * obra.porcentaje / 100)
        except:
            pass
    return JsonResponse({"precio": precio})


# ============================================================
# 📤 Carga del Nomenclador desde Excel
# ============================================================

def parse_price(value):
    """Convierte distintos formatos numéricos en Decimal con 2 decimales."""
    if value is None:
        return Decimal("0.00")

    if isinstance(value, (int, float, Decimal)):
        try:
            return Decimal(str(value)).quantize(Decimal("0.01"))
        except (InvalidOperation, ValueError):
            return Decimal("0.00")

    s = str(value).strip()
    if s in ("", "-"):
        return Decimal("0.00")

    if ',' in s and '.' in s:
        s = s.replace('.', '').replace(',', '.')
    elif ',' in s:
        s = s.replace(',', '.')
    s = s.replace(' ', '').replace('$', '').replace('ARS', '')

    try:
        return Decimal(s).quantize(Decimal("0.01"))
    except (InvalidOperation, ValueError):
        return Decimal("0.00")


@login_required
def cargar_nomenclador(request):
    """Carga o actualiza prestaciones desde un archivo Excel y expira presupuestos antiguos."""
    resumen = {'creadas': 0, 'actualizadas': 0, 'errores': [], 'expirados': 0}

    if request.method == "POST":
        form = NomencladorUploadForm(request.POST, request.FILES)
        if form.is_valid():
            archivo = form.cleaned_data['archivo']

            try:
                with transaction.atomic():  # ⛔ Bloquea todo el proceso en una única transacción
                    # === PROCESAR NOMENCLADOR ===
                    wb = load_workbook(filename=archivo, read_only=True, data_only=True)
                    ws = wb.active

                    for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
                        if i < 4:
                            continue

                        try:
                            # --- Procesar fila individual ---
                            if all((c is None or (isinstance(c, str) and c.strip() == "")) for c in row):
                                continue

                            codigo = str(row[0]).strip() if len(row) >= 1 and row[0] else None
                            nombre = str(row[1]).strip() if len(row) >= 2 and row[1] else None
                            especialista = parse_price(row[2]) if len(row) >= 3 else None
                            gastos = parse_price(row[5]) if len(row) >= 6 else None

                            if not codigo or not nombre:
                                raise ValueError(f"Fila {i}: falta código o nombre")
                            if especialista is None or gastos is None:
                                raise ValueError(f"Fila {i}: faltan valores de especialista o gastos")

                            prest = Prestacion.objects.filter(codigo__iexact=codigo).first()
                            if not prest and not codigo:
                                prest = Prestacion.objects.filter(nombre__iexact=nombre).first()

                            if prest:
                                prest.nombre = nombre
                                prest.codigo = codigo
                                prest.especialista = especialista
                                prest.gastos = gastos
                                prest.save()
                                resumen['actualizadas'] += 1
                            else:
                                Prestacion.objects.create(
                                    nombre=nombre,
                                    codigo=codigo,
                                    especialista=especialista,
                                    gastos=gastos
                                )
                                resumen['creadas'] += 1

                        except Exception as ex:
                            resumen['errores'].append(f"Fila {i}: {str(ex)}")
                    from datetime import timedelta
                    from django.utils import timezone
                    # === EXPIRAR PRESUPUESTOS ANTIGUOS ===
                    fecha_limite = timezone.now() - timedelta(days=10)

                    presupuestos_a_expirar = Presupuesto.objects.filter(
                        estado__in=["pendiente", "autorizado"],
                        pagos__isnull=True,
                        updated_at__lte=fecha_limite
                    ).distinct()

                    resumen['expirados'] = presupuestos_a_expirar.count()

                    # 🧾 Se actualizan todos dentro de la misma transacción
                    for p in presupuestos_a_expirar:
                        p.estado = "expirado"
                        p.updated_at = timezone.now()
                        p.save()

            except Exception as ex:
                resumen['errores'].append(f"Error general en la transacción: {str(ex)}")
                messages.error(request, "Ocurrió un error al procesar el archivo. No se realizaron cambios.")
                return render(request, 'presupuestos/cargar_nomenclador.html', {
                    'form': NomencladorUploadForm(),
                    'resumen': resumen
                })

            # ✅ Solo se llega aquí si todo fue exitoso
            messages.success(
                request,
                f"Nomenclador procesado correctamente. "
                f"Creadas: {resumen['creadas']}, Actualizadas: {resumen['actualizadas']}, "
                f"Expiradas: {resumen['expirados']}."
            )
            return render(request, 'presupuestos/cargar_nomenclador.html', {
                'form': NomencladorUploadForm(),
                'resumen': resumen
            })

    else:
        form = NomencladorUploadForm()

    return render(request, 'presupuestos/cargar_nomenclador.html', {
        'form': form,
        'resumen': resumen
    })




@login_required
def codigos_particulares(request):
    error = None
    mensaje = None

    if request.method == "POST":
        id_edit = request.POST.get("id_edit")
        codigo = request.POST.get("codigo", "").strip()
        nombre = request.POST.get("nombre", "").strip()
        gastos = request.POST.get("gastos") or 0
        especialista = request.POST.get("especialista") or 0

        # Validación: código duplicado si es nuevo
        if not id_edit and Prestacion.objects.filter(codigo=codigo).exists():
            error = f"El código {codigo} ya existe."
        else:
            if id_edit:
                prest = get_object_or_404(Prestacion, id=id_edit)
                prest.nombre = nombre
                prest.gastos = gastos
                prest.especialista = especialista
                prest.save()
                mensaje = "Prestación actualizada correctamente."
            else:
                Prestacion.objects.create(
                    nombre=nombre,
                    codigo=codigo,
                    gastos=gastos,
                    especialista=especialista,
                    tipo="particular",
                    user_made=request.user,
                    user_updated=request.user,
                )
                mensaje = "Prestación creada correctamente."

    prestaciones = Prestacion.objects.filter(tipo="particular").order_by("nombre")
    return render(request, "presupuestos/codigos_particulares.html", {
        "prestaciones": prestaciones,
        "error": error,
        "mensaje": mensaje
    })


@login_required
def eliminar_codigo_particular(request, pk):
    prest = get_object_or_404(Prestacion, pk=pk, tipo="particular")
    if request.method == "POST":
        prest.delete()
    return redirect("presupuestos_app:codigos_particulares")
# ============================================================
# 🔍 Buscar Nomenclador (para autocompletar)
# ============================================================
@login_required
def buscar_nomenclador(request):
    q = request.GET.get('q', '').strip()
    if not q:
        return JsonResponse([], safe=False)
    resultados = (
        Prestacion.objects.filter(
            Q(nombre__icontains=q) | Q(codigo__icontains=q))[:20]
        .values('codigo', 'nombre','gastos','especialista')
    )
    return JsonResponse(list(resultados), safe=False)


 
def format_num(n):
    """Formatea número con punto de miles y coma decimal: 12345.67 → 12.345,67"""
    return f"{n:,.2f}".replace(",", "X").replace(".", ",").replace("X", ".")
@login_required
def imprimir_presupuesto(request, pk):
    presupuesto = Presupuesto.objects.get(pk=pk) 
    if presupuesto.estado == 'pendiente' or presupuesto.estado=='expirado':
        return HttpResponse(
            "<h3 style='color:red; text-align:center; margin-top:50px;'>"
            "❌ No se puede imprimir un presupuesto pendiente."
            "</h3>",
            content_type="text/html"
        )
    items= (
        presupuesto.items
        .annotate(
            prioridad=Case(
                When(codigo__regex=r'^\d+$', then=Value(0)),            # si el código es numérico
                When(codigo__istartswith="med", then=Value(2)),         # si empieza con "med"
                default=Value(1),                                       # el resto
                output_field=IntegerField(),
            )
        )
        .order_by("prioridad", "codigo")  # ordena primero por prioridad, luego por código
    ) 

  

    # --- Crear PDF ---
    response = HttpResponse(content_type='application/pdf')
    filename = f"presupuesto_{pk}_{presupuesto.paciente_nombre}.pdf"
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    c = canvas.Canvas(response, pagesize=A4)
    c.setTitle(filename)  # 👈 Esto define el nombre que ve el navegador
    width, height = A4

    # --- Márgenes ---
    margen_izq = 20*mm
    margen_der = 20*mm
    margen_sup = height - 20*mm
    y_actual = margen_sup

    styles = getSampleStyleSheet()
    
    # --- Encabezado ---
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margen_izq, y_actual, "Hospital Privado Santa Clara de Asís S.A.")
    c.setFont("Helvetica", 9)
    y_actual -= 6*mm
    c.drawString(margen_izq, y_actual, "Urquiza 964 - 4400 Salta")
    y_actual -= 10*mm  # Separación antes de datos generales


            # Dibuja el logo a la derecha, alineado con el título
    import os
    from django.conf import settings
    logo_path = os.path.join(settings.BASE_DIR, "static", "fotos", "hpsca_logo.jpg")

    logo_width = 29 * mm   # ancho del logo
    logo_height = 29 * mm  # alto del logo
    try:
        logo = ImageReader(logo_path)
        c.drawImage(
            logo,
            width - logo_width - margen_der,  # 👈 margen derecho real
            A4[1] - logo_height - 10 * mm,  # posición Y: ajustá este valor fino
            width=logo_width,
            height=logo_height,
            preserveAspectRatio=True,
            mask='auto'
        )
    except Exception as e:
        print("No se pudo cargar el logo:", e)

    c.setFont("Helvetica-Bold", 11) 
    texto = f"PRESUPUESTO Nº {presupuesto.id}"
    ancho_pagina = A4[0]

    # Calcular ancho del texto para centrarlo
    ancho_texto = c.stringWidth(texto, "Helvetica-Bold", 11)
    x_centrado = (ancho_pagina - ancho_texto) / 2

    
    # Dibujar texto centrado
    c.drawString(x_centrado, y_actual, texto)

    # Dibujar línea subrayada debajo del texto
    y_linea = y_actual - 2  # 2 puntos por debajo del texto
    c.setLineWidth(0.6)
    c.line(x_centrado, y_linea, x_centrado + ancho_texto, y_linea)

    # Actualizar posición para lo siguiente
    y_actual -= 10 * mm


    # --- Datos generales ---
    c.setFont("Helvetica", 9)
    c.drawString(margen_izq, y_actual, f"Fecha: {presupuesto.fecha_creacion.strftime('%d/%m/%Y')}")
    c.drawString(width/2, y_actual, f"Obra Social: {presupuesto.obra_social}")
    y_actual -= 6*mm
    c.drawString(margen_izq, y_actual, f"Paciente: {presupuesto.paciente_nombre}")
    c.drawString(width/2, y_actual, f"Edad: {presupuesto.paciente_edad or '  '}                 DNI: {presupuesto.paciente_dni or ''}")
    y_actual -= 6*mm
    c.drawString(margen_izq, y_actual, f"Dirección: {presupuesto.paciente_direccion or ''}")
    c.drawString(width/2, y_actual, f"Teléfono: {presupuesto.paciente_telefono or ''}")
    y_actual -= 6*mm
    c.drawString(margen_izq, y_actual, f"Médico: {presupuesto.medico or ''}")
    c.drawString(width/2, y_actual, f"Diagnóstico: {presupuesto.diagnostico or ''}")
    
    y_actual -= 10 * mm

    # --- Tabla de items ---
    data = [["Código", "Prestación", "Cant.", "P. Unitario", "Importe", "IVA", "Subtotal"]]
    for item in items:
        data.append([
            item.codigo or "",
            item.prestacion.capitalize() or "",
            item.cantidad,
            format_num(item.precio),
            format_num(item.importe),
            format_num(item.iva),
            format_num(item.subtotal),
        ])

    colWidths = [15*mm, 77*mm, 10*mm, 17*mm, 17*mm, 17*mm, 17*mm]
    table = Table(data, colWidths=colWidths, repeatRows=1)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), colors.lightgrey),
        ('GRID', (0,0), (-1,-1), 0.25, colors.grey),
        ('FONT', (0,0), (-1,0), 'Helvetica-Bold',8),
        ('FONT', (0,1), (-1,-1), 'Helvetica', 7),
        ('ALIGN', (2,0), (-1,-1), 'RIGHT'),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
    ]))

    # ✅ Calcular la altura exacta de la tabla
    table_width, table_height = table.wrap(width - margen_izq - margen_der, y_actual)

    # Dibujar la tabla justo debajo del texto anterior
    table.drawOn(c, margen_izq, y_actual - table_height)

    # Actualizar y_actual después de dibujar la tabla
    y_actual -= table_height + 10*mm
        # --- Totales ---
    # --- Totales y Pagos (lado a lado) ---

    c.setFont("Helvetica-Bold", 9)

    # === Columna izquierda: pagos parciales ===
    x_pagos = margen_izq
    y_pagos = y_actual

    c.drawString(x_pagos, y_pagos+2*mm, "Pagos Parciales:")
    c.setFont("Helvetica", 8)
    y_pagos -= 5*mm

    # 🔹 ejemplo de pagos (puedes reemplazar por tus datos reales)
    pagos = [
        (p.fecha.strftime("%d/%m/%Y"), float(p.monto))
        for p in presupuesto.pagos.all().order_by("fecha")
    ]
    for fecha, monto in pagos:
        c.drawString(x_pagos , y_pagos, f"{fecha}:")
        c.drawRightString(width/2 - 6*mm, y_pagos, f"$ {format_num(monto)}")
        y_pagos -= 4*mm

    # Calcular total de pagos y saldo pendiente
    total_pagos = sum((p.monto for p in presupuesto.pagos.all()), Decimal("0"))
    saldo = presupuesto.total - total_pagos

    # Línea divisoria visual entre columnas
    c.setStrokeColor(colors.lightgrey)
    c.setLineWidth(0.4)
    c.line(width/2, y_actual + 2*mm, width/2, y_actual - 20*mm)  # línea vertical separadora

    # === Columna derecha: totales ===
    x_totales = width - margen_der
    y_totales = y_actual

    c.setFont("Helvetica-Bold", 9)
    c.drawRightString(x_totales, y_totales, f"Subtotal (sin IVA): $ {format_num(presupuesto.subtotal)}")
    y_totales -= 6*mm
    c.drawRightString(x_totales, y_totales, f"IVA (21%): $ {format_num(presupuesto.iva)}")
    y_totales -= 6*mm
    c.drawRightString(x_totales, y_totales, f"Total Presupuesto: $ {format_num(presupuesto.total)}")

    # 💰 Saldo pendiente destacado
    y_totales -= 8*mm
    c.setFont("Helvetica-Bold", 10) 
    c.drawRightString(x_totales, y_totales, f"Saldo pendiente: $ {format_num(saldo)}")
    c.setFillColor(colors.black)

    # --- Observaciones ---
    y_actual = y_totales - 8*mm
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margen_izq, y_actual, "OBSERVACIONES:")
    c.setFont("Helvetica", 9)
    y_actual -= 5*mm
    texto_obs = """\
El monto presupuestado es ESTIMATIVO.    
El presente presupuesto tiene una validez de 7 días habiles.
El presupuesto debe estar abonado en forma previa al ingreso del paciente.

NO INCLUYE Honorarios Médicos.
NO INCLUYE Honorarios Anestesistas (Dirigirse a Sra. Viviana Baigorria 3872259689)
NO INCLUYE Honorarios por Transfusiones de Sangre. En caso de necesitar, dirigirse a BANCO DE SANGRE.
NO INCLUYE prácticas no detalladas.
NO INCLUYE medicamentos de alto costo.

Puede consultar su liquidación final 72hs habiles posteriores al alta."""
    text_obj = c.beginText(margen_izq, y_actual)
    text_obj.setFont("Helvetica", 8)

    for line in texto_obs.split("\n"):
        if line.strip() == "":
            # Dibuja una línea horizontal corta en lugar del salto vacío
            c.drawString(margen_izq + 20, y_actual, "" * 20)  # Línea visual corta
            y_actual -= 0.8 * mm  # Pequeño espacio debajo
        else:
            text_obj.setTextOrigin(margen_izq, y_actual)
            text_obj.textLine(line)
            y_actual -= 4 * mm
    
    c.drawText(text_obj)
    y_actual -= 4 * mm

    # --- Formas de pago ---
    c.setFont("Helvetica-Bold", 10)
    c.drawString(margen_izq, y_actual, "FORMAS DE PAGO:")
    y_actual -= 6*mm
    c.setFont("Helvetica", 9)
    c.drawString(margen_izq, y_actual, "Efectivo, tarjeta de crédito, débito o transferencia (CBU: 2850100630000800105391 ALIAS:SOGA.BOLSA.COBRE)")
    y_actual -= 5*mm
    c.drawString(margen_izq, y_actual, "Hospital Privado Santa Clara de Asís S.A.")
    y_actual -= 25*mm
    # --- Firma ---
    from datetime import datetime
    # --- Fecha y hora de impresión ---
    fecha_impresion = datetime.now().strftime("%d/%m/%Y %H:%M")
    c.setFont("Helvetica", 9)
    c.drawString(margen_izq, y_actual + 5*mm, f"Fecha y hora de impresión: {fecha_impresion}")


    c.drawString(margen_izq, y_actual, f"Confeccionó: {presupuesto.user_updated.last_name}, {presupuesto.user_updated.first_name}")
    # --- Firma digital autorizada ---
    firma_path = os.path.join(settings.BASE_DIR, "static", "fotos", "paolaulloa.jpg")
    firma_width = 55 * mm   # ancho deseado de la firma
    firma_height = 25 * mm  # alto deseado de la firma

    try:
        firma_img = ImageReader(firma_path)
        c.drawImage(
            firma_img,
            width - margen_der - 53*mm,  # mismo eje X que la línea de firma
            y_actual - 1*mm,             # un poco arriba de la línea
            width=firma_width,
            height=firma_height,
            preserveAspectRatio=True,
            mask='auto'
        )
    except Exception as e:
        print("No se pudo cargar la firma digital:", e)
    c.drawString(width-margen_der-50*mm, y_actual, "_________________________")
    c.drawString(width-margen_der-40*mm, y_actual-6*mm, "Firma Autorizado")

    c.showPage()
    c.save()
    return response




from django.shortcuts import render
from django.db.models import Q
from django.http import HttpResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Presupuesto


def reporte_presupuestos_fecha(request):
    presupuestos = Presupuesto.objects.all().select_related(
        "user_made",
        "medico",
    ).order_by("-updated_at")

    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    estado = request.GET.get("estado", "").strip()

    if fecha_desde:
        presupuestos = presupuestos.filter(updated_at__date__gte=fecha_desde)

    if fecha_hasta:
        presupuestos = presupuestos.filter(updated_at__date__lte=fecha_hasta)

    if estado:
        presupuestos = presupuestos.filter(estado=estado)

    # Exportar Excel
    if request.GET.get("exportar") == "excel":
        return exportar_presupuestos_excel(
            presupuestos=presupuestos,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            estado=estado,
        )

    context = {
        "presupuestos": presupuestos,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "estado": estado,
        "estados": [
            ("pendiente", "Pendiente"),
            ("autorizado", "Autorizado"),
            ("expirado", "Expirado"),
            ("rechazado", "Rechazado"),
        ],
        "total_registros": presupuestos.count(),
    }
    return render(request, "presupuestos/reporte_presupuestos_fecha.html", context)


def exportar_presupuestos_excel(presupuestos, fecha_desde="", fecha_hasta="", estado=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Presupuestos"

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"] = "Reporte de Presupuestos por Fecha"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0D6EFD")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Subfiltros
    ws.merge_cells("A2:E2")
    filtros_texto = f"Desde: {fecha_desde or '-'} | Hasta: {fecha_hasta or '-'} | Estado: {estado or 'Todos'}"
    ws["A2"] = filtros_texto
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Cabeceras
    headers = ["ID Presupuesto", "Fecha", "Paciente", "Médico", "Estado"]
    header_row = 4

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="198754")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    row = header_row + 1
    for p in presupuestos:
        paciente = getattr(p, "paciente", None)
        medico = getattr(p, "medico", None)

        ws.cell(row=row, column=1, value=p.id)
        ws.cell(
            row=row,
            column=2,
            value=p.fecha.strftime("%d/%m/%Y %H:%M") if p.fecha else ""
        )
        ws.cell(row=row, column=3, value=str(paciente) if paciente else "-")
        ws.cell(row=row, column=4, value=str(medico) if medico else "-")
        ws.cell(row=row, column=5, value=p.estado or "-")
        row += 1

    # Anchos
    widths = {
        1: 18,
        2: 22,
        3: 30,
        4: 30,
        5: 18,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # Congelar cabecera
    ws.freeze_panes = "A5"

    filename = "reporte_presupuestos.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response




def reporte_pagos_fecha(request):
    pagos = Pago.objects.select_related(
        "presupuesto",
        "presupuesto__medico",
        "user_made",
    ).order_by("-fecha")

    fecha_desde = request.GET.get("fecha_desde", "").strip()
    fecha_hasta = request.GET.get("fecha_hasta", "").strip()
    caja = request.GET.get("caja", "").strip()

    if fecha_desde:
        pagos = pagos.filter(fecha__date__gte=fecha_desde)

    if fecha_hasta:
        pagos = pagos.filter(fecha__date__lte=fecha_hasta)

    if caja:
        pagos = pagos.filter(caja=caja)

    if request.GET.get("exportar") == "excel":
        return exportar_pagos_excel(
            pagos=pagos,
            fecha_desde=fecha_desde,
            fecha_hasta=fecha_hasta,
            caja=caja,
        )

    context = {
        "pagos": pagos,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "caja": caja,
        "cajas": Pago.TIPO_CAJA,
        "total_registros": pagos.count(),
        "total_importe": sum(p.monto for p in pagos),
    }
    return render(request, "presupuestos/reporte_pagos_fecha.html", context)


def exportar_pagos_excel(pagos, fecha_desde="", fecha_hasta="", caja=""):
    wb = Workbook()
    ws = wb.active
    ws.title = "Reporte Pagos"

    # Título
    ws.merge_cells("A1:E1")
    ws["A1"] = "Reporte de Pagos por Fecha"
    ws["A1"].font = Font(bold=True, size=14, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor="0D6EFD")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")

    # Filtros
    ws.merge_cells("A2:E2")
    filtros_texto = f"Desde: {fecha_desde or '-'} | Hasta: {fecha_hasta or '-'} | Caja: {caja or 'Todas'}"
    ws["A2"] = filtros_texto
    ws["A2"].font = Font(italic=True)
    ws["A2"].alignment = Alignment(horizontal="center")

    # Cabeceras
    headers = ["Importe", "Fecha", "Caja", "Paciente", "Medio de Pago"]
    header_row = 4

    for col_num, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col_num, value=header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="198754")
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # Datos
    row = header_row + 1
    total_importe = 0

    for pago in pagos:
        paciente = pago.presupuesto.paciente_nombre if pago.presupuesto else "-"
        caja_display = pago.get_caja_display() if pago.caja else "-"
        medio_pago_display = pago.get_medio_pago_display() if pago.medio_pago else "-"

        ws.cell(row=row, column=1, value=float(pago.monto))
        ws.cell(row=row, column=2, value=pago.fecha.strftime("%d/%m/%Y %H:%M") if pago.fecha else "")
        ws.cell(row=row, column=3, value=caja_display)
        ws.cell(row=row, column=4, value=paciente)
        ws.cell(row=row, column=5, value=medio_pago_display)

        total_importe += pago.monto
        row += 1

    # Fila total
    ws.cell(row=row + 1, column=1, value="TOTAL")
    ws.cell(row=row + 1, column=1).font = Font(bold=True)
    ws.cell(row=row + 1, column=2, value=float(total_importe))
    ws.cell(row=row + 1, column=2).font = Font(bold=True)

    # Formato moneda columna importe
    for r in range(header_row + 1, row + 2):
        ws.cell(row=r, column=1).number_format = '$ #,##0.00'

    # Anchos
    widths = {
        1: 18,
        2: 22,
        3: 18,
        4: 35,
        5: 22,
    }

    for col_idx, width in widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.freeze_panes = "A5"

    filename = "reporte_pagos.xlsx"
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f'attachment; filename="{filename}"'
    wb.save(response)
    return response





#reportes 


from decimal import Decimal
from collections import defaultdict

from django.contrib.auth.decorators import login_required
from django.db.models import Count, Sum, Q
from django.db.models.functions import TruncMonth
from django.shortcuts import render

from .models import Presupuesto, Pago, Reintegro, Medico, ObraSocial

from datetime import date

@login_required
def reporte_resumen_general(request):
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    medico_id = request.GET.get("medico")
    obra_social_id = request.GET.get("obra_social")
    estado = request.GET.get("estado")

    presupuestos = (
        Presupuesto.objects
        .select_related("medico", "obra_social")
        .prefetch_related("items", "pagos", "reintegros")
        .all()
        .order_by("-fecha_creacion")
    )

    if fecha_desde:
        presupuestos = presupuestos.filter(fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        presupuestos = presupuestos.filter(fecha_creacion__date__lte=fecha_hasta)
    if medico_id:
        presupuestos = presupuestos.filter(medico_id=medico_id)
    if obra_social_id:
        presupuestos = presupuestos.filter(obra_social_id=obra_social_id)
    if estado:
        presupuestos = presupuestos.filter(estado=estado)

    presupuestos_list = list(presupuestos)

    # =========================
    # KPIs GENERALES
    # =========================
    total_presupuestos = len(presupuestos_list)
    total_presupuestado = sum((p.total for p in presupuestos_list), Decimal("0.00"))

    total_cobrado = Decimal("0.00")
    total_reintegrado = Decimal("0.00")
    cantidad_ejecutados = 0
    cantidad_presupuestados = 0  # sin pagos

    for p in presupuestos_list:
        pagos = list(p.pagos.all())
        reintegros = list(p.reintegros.all())

        pagado = sum((pg.monto for pg in pagos), Decimal("0.00"))
        reintegrado = sum((rg.monto for rg in reintegros), Decimal("0.00"))

        total_cobrado += pagado
        total_reintegrado += reintegrado

        if pagos:
            cantidad_ejecutados += 1
        else:
            cantidad_presupuestados += 1

    saldo_global = total_presupuestado - total_cobrado + total_reintegrado
    porcentaje_ejecucion = round((cantidad_ejecutados / total_presupuestos) * 100, 2) if total_presupuestos else 0

    # =========================
    # RESUMEN POR ESTADO
    # =========================
    estados_base = {key: 0 for key, _ in Presupuesto.ESTADOS}
    for p in presupuestos_list:
        estados_base[p.estado] += 1

    estados_resumen = [
        {
            "key": key,
            "label": label,
            "cantidad": estados_base.get(key, 0),
        }
        for key, label in Presupuesto.ESTADOS
    ]

    # =========================
    # TOP MÉDICOS
    # =========================
    medicos_map = defaultdict(lambda: {
        "medico": "",
        "cantidad": 0,
        "ejecutados": 0,
        "total": Decimal("0.00"),
        "cobrado": Decimal("0.00"),
        "saldo": Decimal("0.00"),
    })

    for p in presupuestos_list:
        pagos = list(p.pagos.all())
        pagado = sum((pg.monto for pg in pagos), Decimal("0.00"))

        med_id = p.medico_id
        medicos_map[med_id]["medico"] = str(p.medico)
        medicos_map[med_id]["cantidad"] += 1
        medicos_map[med_id]["total"] += p.total
        medicos_map[med_id]["cobrado"] += pagado
        medicos_map[med_id]["saldo"] += p.saldo

        if pagos:
            medicos_map[med_id]["ejecutados"] += 1

    resumen_medicos = sorted(
        medicos_map.values(),
        key=lambda x: (x["total"], x["cantidad"]),
        reverse=True
    )[:8]

    # =========================
    # CHART: ÚLTIMOS 8 MESES
    # Presupuestado = sin pagos
    # Ejecutado = con algún pago
    # =========================
    def month_start(dt):
        return date(dt.year, dt.month, 1)

    def previous_month(dt):
        if dt.month == 1:
            return date(dt.year - 1, 12, 1)
        return date(dt.year, dt.month - 1, 1)

    # referencia final
    if presupuestos_list:
        fecha_referencia = max(p.fecha_creacion.date() for p in presupuestos_list if p.fecha_creacion)
    else:
        from django.utils import timezone
        fecha_referencia = timezone.localdate()

    # si hay fecha_hasta, usamos ese corte
    if fecha_hasta:
        try:
            y, m, d = map(int, fecha_hasta.split("-"))
            fecha_referencia = date(y, m, d)
        except Exception:
            pass

    fecha_referencia = month_start(fecha_referencia)

    # generamos hasta 8 meses hacia atrás
    meses = []
    cursor = fecha_referencia
    for _ in range(8):
        meses.append(cursor)
        cursor = previous_month(cursor)
    meses.reverse()

    # si hay fecha_desde, quitamos meses anteriores al filtro
    if fecha_desde:
        try:
            y, m, d = map(int, fecha_desde.split("-"))
            desde_mes = date(y, m, 1)
            meses = [m for m in meses if m >= desde_mes]
        except Exception:
            pass

    chart_map = {
        m.strftime("%Y-%m"): {
            "label": m.strftime("%m/%Y"),
            "presupuestado": 0,  # sin pagos
            "ejecutado": 0,      # con pagos
        }
        for m in meses
    }

    for p in presupuestos_list:
        if not p.fecha_creacion:
            continue

        clave = p.fecha_creacion.strftime("%Y-%m")
        if clave not in chart_map:
            continue

        tiene_pago = p.pagos.exists()
        if tiene_pago:
            chart_map[clave]["ejecutado"] += 1
        else:
            chart_map[clave]["presupuestado"] += 1

    chart_data = [chart_map[m.strftime("%Y-%m")] for m in meses]

    chart_labels = [item["label"] for item in chart_data]
    chart_presupuestado = [item["presupuestado"] for item in chart_data]
    chart_ejecutado = [item["ejecutado"] for item in chart_data]

    # =========================
    # ÚLTIMOS PRESUPUESTOS
    # =========================
    ultimos_presupuestos = presupuestos_list[:10]

    context = {
        "titulo": "Resumen General",
        "total_presupuestos": total_presupuestos,
        "total_presupuestado": total_presupuestado,
        "total_cobrado": total_cobrado,
        "total_reintegrado": total_reintegrado,
        "saldo_global": saldo_global,
        "cantidad_ejecutados": cantidad_ejecutados,
        "cantidad_presupuestados": cantidad_presupuestados,
        "porcentaje_ejecucion": porcentaje_ejecucion,
        "estados_resumen": estados_resumen,
        "resumen_medicos": resumen_medicos,
        "ultimos_presupuestos": ultimos_presupuestos,
        "medicos": Medico.objects.all().order_by("nombre"),
        "obras_sociales": ObraSocial.objects.all().order_by("nombre"),
        "estado_actual": estado,
        "medico_actual": medico_id,
        "obra_social_actual": obra_social_id,
        "fecha_desde": fecha_desde,
        "fecha_hasta": fecha_hasta,
        "estados": Presupuesto.ESTADOS,
        "chart_labels_json": json.dumps(chart_labels),
        "chart_presupuestado_json": json.dumps(chart_presupuestado),
        "chart_ejecutado_json": json.dumps(chart_ejecutado),
    }
    return render(request, "presupuestos/reportes/resumen_general.html", context)






@login_required
def reporte_presupuestos(request):
    fecha_desde = request.GET.get("fecha_desde")
    fecha_hasta = request.GET.get("fecha_hasta")
    paciente = request.GET.get("paciente")
    dni = request.GET.get("dni")
    medico_id = request.GET.get("medico")
    obra_social_id = request.GET.get("obra_social")
    estado = request.GET.get("estado")
    solo_ejecutados = request.GET.get("solo_ejecutados")
    exportar = request.GET.get("exportar")

    qs = (
        Presupuesto.objects
        .select_related("medico", "obra_social")
        .prefetch_related("pagos", "reintegros", "items")
        .order_by("-fecha_creacion")
    )

    # FILTROS
    if fecha_desde:
        qs = qs.filter(fecha_creacion__date__gte=fecha_desde)
    if fecha_hasta:
        qs = qs.filter(fecha_creacion__date__lte=fecha_hasta)
    if paciente:
        qs = qs.filter(paciente_nombre__icontains=paciente)
    if dni:
        qs = qs.filter(paciente_dni__icontains=dni)
    if medico_id:
        qs = qs.filter(medico_id=medico_id)
    if obra_social_id:
        qs = qs.filter(obra_social_id=obra_social_id)
    if estado:
        qs = qs.filter(estado=estado)

    presupuestos = list(qs)

    # CALCULOS
    total_presupuestado = Decimal("0")
    total_cobrado = Decimal("0")
    total_reintegrado = Decimal("0")

    data = []

    for p in presupuestos:
        pagos = list(p.pagos.all())
        reintegros = list(p.reintegros.all())

        cobrado = sum((pg.monto for pg in pagos), Decimal("0"))
        reintegrado = sum((rg.monto for rg in reintegros), Decimal("0"))
        saldo = p.total - cobrado + reintegrado

        ejecutado = True if pagos else False

        if solo_ejecutados == "1" and not ejecutado:
            continue

        total_presupuestado += p.total
        total_cobrado += cobrado
        total_reintegrado += reintegrado

        data.append({
            "obj": p,
            "cobrado": cobrado,
            "reintegrado": reintegrado,
            "saldo": saldo,
            "ejecutado": ejecutado
        })

    # =========================
    # EXPORT EXCEL
    # =========================
    if exportar == "excel":
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Presupuestos"

        headers = [
            "Nro", "Fecha", "Paciente", "DNI",
            "Medico", "Obra Social", "Estado",
            "Total", "Cobrado", "Reintegrado", "Saldo"
        ]

        ws.append(headers)

        for item in data:
            p = item["obj"]
            ws.append([
                p.id,
                p.fecha_creacion.strftime("%d/%m/%Y"),
                p.paciente_nombre,
                p.paciente_dni,
                str(p.medico),
                str(p.obra_social),
                p.get_estado_display(),
                float(p.total),
                float(item["cobrado"]),
                float(item["reintegrado"]),
                float(item["saldo"]),
            ])

        response = HttpResponse(
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )
        response["Content-Disposition"] = "attachment; filename=presupuestos.xlsx"
        wb.save(response)
        return response

    context = {
        "data": data,
        "total_presupuestado": total_presupuestado,
        "total_cobrado": total_cobrado,
        "total_reintegrado": total_reintegrado,
        "saldo_total": total_presupuestado - total_cobrado + total_reintegrado,
        "medicos": Medico.objects.all(),
        "obras_sociales": ObraSocial.objects.all(),
        "estados": Presupuesto.ESTADOS,
        "request": request
    }

    return render(request, "presupuestos/reportes/presupuestos_detallado.html", context)